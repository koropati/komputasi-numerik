from openpyxl import load_workbook

class ExcelReader:
    def __init__(self, file_path, sheet_name=None):
        """
        Inisialisasi ExcelReader
        :param file_path: path ke file Excel yang ingin dibaca
        :param sheet_name: nama sheet yang ingin dibaca (default = None untuk sheet pertama)
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None

    def load_excel(self):
        """
        Membuka file Excel dan memuat sheet yang ditentukan
        """
        try:
            self.workbook = load_workbook(filename=self.file_path)
            if self.sheet_name:
                self.sheet = self.workbook[self.sheet_name]
            else:
                self.sheet = self.workbook.active  # Ambil sheet pertama jika tidak ada nama sheet
            print(f"File Excel {self.file_path} berhasil dibuka.")
        except Exception as e:
            print(f"Error saat membuka file Excel: {e}")
            return None

    def get_cholesky_data(self, start_row=1, start_col=1, end_row=None, end_col=None):
        """
        Mengambil data dalam bentuk array dari sheet berdasarkan range yang diberikan
        :param start_row: baris awal data (default = 1)
        :param start_col: kolom awal data (default = 1)
        :param end_row: baris akhir data (default = None, sampai akhir)
        :param end_col: kolom akhir data (default = None, sampai akhir)
        :return: array data yang diambil dari Excel
        """
        if self.sheet is None:
            print("Belum ada sheet yang dibaca. Gunakan load_excel() terlebih dahulu.")
            return None
        
        # Tentukan batas baris dan kolom jika tidak diberikan
        end_row = end_row or self.sheet.max_row
        end_col = end_col or self.sheet.max_column

        # Ambil data dari rentang yang diberikan
        data = []
        for row in self.sheet.iter_rows(min_row=start_row, max_row=end_row, 
                                        min_col=start_col, max_col=end_col, values_only=True):
            data.append(list(row))
        
        print("Data array berhasil diambil.")
        return data